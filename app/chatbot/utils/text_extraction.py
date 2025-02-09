import io
import csv
import base64
import regex as re
from docx import Document
from openpyxl import load_workbook
from fastapi import UploadFile
import fitz  # PyMuPDF
from .openai import get_openai_client
from .structlogger import logger
import os
from datetime import datetime
import asyncio
import time
from asyncio import Semaphore
import aiofiles

# Constants for rate limiting and failure handling
MAX_CONCURRENT_REQUESTS = 5
REQUESTS_PER_MINUTE = 30  # Increased to 30 requests per minute
REQUEST_INTERVAL = 2.0  # 2 seconds between requests
LLM_FAILURE_LIMIT = 18
LLM_FAILURE_COUNT = 0
INITIAL_DELAY = 1

# Create a semaphore to limit concurrent requests
api_semaphore = Semaphore(MAX_CONCURRENT_REQUESTS)


class RateLimiter:
    def __init__(self, max_calls: int, interval: float):
        self.max_calls = max_calls
        self.interval = interval
        self.last_request_time = 0.0
        self.lock = asyncio.Lock()

    async def __aenter__(self):
        async with self.lock:
            now = time.time()
            time_since_last = now - self.last_request_time
            if time_since_last < self.interval:
                await asyncio.sleep(self.interval - time_since_last)
            self.last_request_time = time.time()

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        pass


# Update rate limiter instance
rate_limiter = RateLimiter(REQUESTS_PER_MINUTE, REQUEST_INTERVAL)


async def gpt4v_process_image(client, messages):
    """
    Process image with GPT-4V with rate limiting and failure tracking
    """
    global LLM_FAILURE_COUNT
    retries = 0
    max_retries = 6

    while retries < max_retries:
        try:
            async with api_semaphore, rate_limiter:
                response = await client.chat.completions.create(
                    model="gpt-4o",
                    messages=messages,
                    temperature=0.0,
                )
                LLM_FAILURE_COUNT = 0  # Reset failure count on success
                return response

        except Exception as e:
            retries += 1
            LLM_FAILURE_COUNT += 1
            logger.warning(f"GPT-4V processing attempt {retries} failed: {str(e)}")

            if LLM_FAILURE_COUNT >= LLM_FAILURE_LIMIT:
                logger.critical("Too many consecutive LLM failures. Stopping.")
                raise Exception("Too many consecutive LLM failures")

            if retries % 3 == 0:
                logger.info("Waiting for 10 seconds before retrying...")
                await asyncio.sleep(10)
            else:
                await asyncio.sleep(INITIAL_DELAY)

    logger.error("Max retries reached for GPT-4V processing")
    raise Exception("Max retries reached for GPT-4V processing")


async def extract_text_from_txt(file: UploadFile) -> str:
    content = await file.read()
    return content.decode("utf-8")


async def extract_text_from_pdf(file: UploadFile) -> str:
    """
    Extract text from PDF with fallback to OCR using GPT-4V for image-based PDFs
    """
    content = await file.read()
    pdf = fitz.open(stream=content, filetype="pdf")
    total_pages = len(pdf)

    # First attempt: Direct text extraction
    text = ""
    for page in pdf:
        text += page.get_text()

    # Check if meaningful text was extracted using multiple criteria
    text_stripped = text.strip()
    if is_text_meaningful(text_stripped, total_pages):
        logger.info("Successfully extracted meaningful text directly from PDF")
        return text

    logger.info(
        "Direct text extraction failed quality checks, attempting OCR with GPT-4V"
    )

    # Fallback: Convert pages to images and use GPT-4V
    try:
        client = await get_openai_client()
        # Process all pages concurrently
        tasks = []
        for page_num in range(len(pdf)):
            page = pdf[page_num]
            base64_image = await convert_page_to_base64(page)
            task = process_page_with_gpt4v(client, base64_image, page_num, len(pdf))
            tasks.append(task)

        # Wait for all pages to be processed
        page_texts = await asyncio.gather(*tasks)
        combined_text = "\n".join(page_texts)

        # Save debug output asynchronously
        await save_debug_output(text_stripped, combined_text)

        logger.info("Successfully extracted text using GPT-4V OCR")
        return combined_text.strip()

    except Exception as e:
        logger.error(f"Error during GPT-4V OCR processing: {str(e)}")
        return text_stripped


def is_text_meaningful(text: str, total_pages: int = 0) -> bool:
    """
    Enhanced check for text quality, including detection of tables and figures
    Forces OCR for documents <= 10 pages or with potential tables/figures
    """
    # Patterns to detect various text elements
    patterns = {
        # Basic text patterns
        "sentences": r"[A-Z][^.!?]*[.!?]",
        "words": r"\b\w{2,}\b",
        
        # Table detection patterns
        "column_headers": r"(?:\s{2,}|\t)[A-Z][a-zA-Z\s]*(?:\s{2,}|\t)",
        "data_rows": r"\d+[\s\t]+[A-Za-z]+[\s\t]+\d+",
        "table_borders": r"[+\-|]{3,}",
        "aligned_text": r"(?m)^[\s\t]*\w+[\s\t]{2,}\w+",
        "tabular_spacing": r"(?:\s{2,}|\t)\S+(?:\s{2,}|\t)\S+",
        
        # Figure detection patterns
        "figure_captions": r"(?i)(?:figure|fig\.|table|tbl\.)[\s\d\.:]+\w+",
        "image_placeholders": r"\[(?:image|figure|graph|chart|diagram|photo)\]",
        "reference_markers": r"\[[\d,\s-]+\]|\(\d+\)",
        
        # Layout artifacts
        "page_numbers": r"\b\d+\s*(?:of|\/)?\s*\d+\b",
        "headers_footers": r"(?m)^.{1,20}$[\r\n]+(?:={3,}|-{3,})",
        "bullet_points": r"(?m)^[\s]*[•\-\*][\s]+\w+",
    }
    
    # Force OCR for small documents
    if 0 < total_pages <= 10:
        logger.info("Document is <= 10 pages, forcing OCR processing")
        return False
    
    # Check for basic text quality
    has_basic_text = (
        len(text) > 50
        and bool(re.search(patterns["sentences"], text))
        and len(re.findall(patterns["words"], text)) > 10
    )
    
    # Check for table indicators
    has_table_indicators = any([
        bool(re.search(patterns["column_headers"], text)),
        bool(re.search(patterns["data_rows"], text)),
        bool(re.search(patterns["table_borders"], text)),
        bool(re.search(patterns["aligned_text"], text, re.MULTILINE)),
        bool(re.search(patterns["tabular_spacing"], text)),
        len(re.findall(r"\t", text)) > 5,  # Multiple tab characters
        bool(re.search(r"(?m)^\s*\d+\s+\w+\s+\d+\s*$", text))  # Typical table row
    ])
    
    # Check for figure and layout indicators
    has_figure_indicators = any([
        bool(re.search(patterns["figure_captions"], text)),
        bool(re.search(patterns["image_placeholders"], text)),
        bool(re.search(patterns["reference_markers"], text)),
        bool(re.search(patterns["headers_footers"], text))
    ])
    
    # Additional checks for suspicious formatting
    has_suspicious_formatting = any([
        text.count('\n\n') > text.count('.') * 0.5,  # Too many double line breaks
        len(re.findall(r"\s{3,}", text)) > 10,  # Multiple spaces
        bool(re.search(r"\|\s*\w+\s*\|", text)),  # Pipe characters (common in tables)
        bool(re.search(r"^\s*[\w\d]+:.*(?:\n\s*[\w\d]+:.*){2,}", text, re.MULTILINE))  # List-like structures
    ])

    logger.info(
        "Text quality check results",
        has_basic_text=has_basic_text,
        has_table_indicators=has_table_indicators,
        has_figure_indicators=has_figure_indicators,
        has_suspicious_formatting=has_suspicious_formatting
    )

    # Return False if we detect tables or figures, forcing OCR processing
    if has_table_indicators or has_figure_indicators or has_suspicious_formatting:
        logger.info("Detected tables, figures, or suspicious formatting - forcing OCR processing")
        return False

    return has_basic_text


async def convert_page_to_base64(page) -> str:
    """Convert PDF page to base64 encoded PNG"""
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    img_data = pix.tobytes("png")
    return base64.b64encode(img_data).decode("utf-8")


async def process_page_with_gpt4v(
    client, base64_image: str, page_num: int, total_pages: int
) -> str:
    """Process a single page with GPT-4V with robust error handling"""
    max_retries = 3
    retry_count = 0

    while retry_count < max_retries:
        try:
            messages = [
                {"role": "system", "content": pdf_extraction_prompt},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "Here is the image to extract text from:",
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{base64_image}",
                                "detail": "high",
                            },
                        },
                    ],
                },
            ]

            response = await gpt4v_process_image(client, messages)
            page_text = response.choices[0].message.content
            logger.info(
                f"Successfully processed page {page_num + 1}/{total_pages} with GPT-4V"
            )
            return page_text

        except Exception as e:
            retry_count += 1
            logger.warning(
                f"Failed to process page {page_num + 1}/{total_pages} (attempt {retry_count}): {str(e)}"
            )
            if retry_count < max_retries:
                await asyncio.sleep(5 * retry_count)  # Exponential backoff
            else:
                logger.error(
                    f"Failed to process page {page_num + 1}/{total_pages} after {max_retries} attempts"
                )
                raise


async def save_debug_output(original_text: str, ocr_text: str):
    """Save debug output to file asynchronously"""
    debug_dir = "debug_extractions"
    os.makedirs(debug_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{debug_dir}/extracted_text_{timestamp}.txt"

    # Find next available filename
    counter = 1
    while os.path.exists(filename):
        filename = f"{debug_dir}/extracted_text_{timestamp}_{counter}.txt"
        counter += 1

    try:
        async with aiofiles.open(filename, "w", encoding="utf-8") as f:
            await f.write(f"Original Extraction Attempt:\n{'='*50}\n")
            await f.write(original_text)
            await f.write(f"\n\nGPT-4V OCR Extraction:\n{'='*50}\n")
            await f.write(ocr_text)
        logger.info(f"Debug output saved to {filename}")
    except Exception as e:
        logger.error(f"Failed to save debug output: {str(e)}")


async def extract_text_from_docx(file: UploadFile) -> str:
    content = await file.read()
    doc = Document(io.BytesIO(content))
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])


async def extract_text_from_excel(file: UploadFile) -> str:
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content))
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += "\t".join([str(cell) for cell in row if cell is not None]) + "\n"
    return text


async def extract_text_from_csv(file: UploadFile) -> str:
    content = await file.read()
    csv_content = content.decode("utf-8").splitlines()
    reader = csv.reader(csv_content)
    return "\n".join(["\t".join(row) for row in reader])


pdf_extraction_prompt = """
Please extract the text from this image in a readable, continuous paragraph format, maintaining the logical grouping of sections, headers, and subheaders as seen in the original layout. Avoid unnecessary line breaks or gaps, ensuring the text flows like an article.

If you detect a table:
- Convert the table into a descriptive paragraph. For each row, use a sentence or phrase that clearly conveys the meaning of the table’s data. For example, "The course EEE325 covers topics such as digital systems and is led by faculty from the Electrical Engineering department." 
- Ensure that each row and column’s information is maintained in a logical, coherent sentence structure, preserving data relationships without the need for grid formatting.
- For tables with complex data, such as mathematical formulas or structured data, describe the content briefly and indicate the presence of complex information for later review.

For any embedded images or diagrams:
- Provide a brief text description if they contain essential information. If they are purely decorative or non-informative, omit them from the output.
"""
